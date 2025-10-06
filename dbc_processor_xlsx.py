"""
Procesador de archivos DBC para extraer información y exportar a Excel
Autor: GitHub Copilot
Fecha: 24 de septiembre de 2025
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from typing import Dict, List, Any, Optional
import logging

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DBCProcessor:
    """
    Clase para procesar archivos DBC y extraer información de mensajes CAN
    """
    
    def __init__(self, dbc_file_path: str):
        """
        Inicializa el procesador DBC
        
        Args:
            dbc_file_path (str): Ruta al archivo DBC
        """
        self.dbc_file_path = dbc_file_path
        self.dbc_content = ""
        self.messages = {}
        self.signals = {}
        self.nodes = []
        self.attributes = {}
        self.value_tables = {}
        self.comments = {}
        
    def load_dbc_file(self) -> bool:
        """
        Carga el archivo DBC en memoria
        
        Returns:
            bool: True si se cargó correctamente, False en caso contrario
        """
        # Intentar diferentes codificaciones
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
        
        for encoding in encodings:
            try:
                with open(self.dbc_file_path, 'r', encoding=encoding) as file:
                    self.dbc_content = file.read()
                logger.info(f"Archivo DBC cargado exitosamente con codificación {encoding}: {self.dbc_file_path}")
                return True
            except (UnicodeDecodeError, UnicodeError):
                logger.debug(f"Falló codificación {encoding}, intentando siguiente...")
                continue
            except FileNotFoundError:
                logger.error(f"Archivo no encontrado: {self.dbc_file_path}")
                return False
            except Exception as e:
                logger.error(f"Error al cargar archivo DBC con {encoding}: {e}")
                continue
        
        logger.error(f"No se pudo cargar el archivo con ninguna codificación probada")
        return False
    
    def parse_nodes(self):
        """
        Extrae los nodos del archivo DBC
        """
        # Buscar declaración de nodos
        nodes_pattern = r'BS_:\s*([^;]+);'
        nodes_match = re.search(nodes_pattern, self.dbc_content)
        
        if nodes_match:
            nodes_str = nodes_match.group(1).strip()
            self.nodes = [node.strip() for node in nodes_str.split() if node.strip()]
        
        logger.info(f"Nodos encontrados: {len(self.nodes)}")
    
    def parse_messages(self):
        """
        Extrae los mensajes del archivo DBC
        """
        # Patrón para mensajes: BO_ <ID> <Name>: <Size> <Node>
        message_pattern = r'BO_\s+(\d+)\s+([^:]+):\s+(\d+)\s+([^\n]+)'
        
        for match in re.finditer(message_pattern, self.dbc_content):
            msg_id = int(match.group(1))
            msg_name = match.group(2).strip()
            msg_size = int(match.group(3))
            msg_node = match.group(4).strip()
            
            self.messages[msg_id] = {
                'id': msg_id,
                'name': msg_name,
                'size': msg_size,
                'node': msg_node,
                'signals': []
            }
        
        logger.info(f"Mensajes encontrados: {len(self.messages)}")
    
    def parse_signals(self):
        """
        Extrae las señales del archivo DBC
        """
        # Patrón para señales: SG_ <Name> : <StartBit>|<Length>@<Endianness><Sign> (<Factor>,<Offset>) [<Min>|<Max>] "<Unit>" [<Nodes>]
        signal_pattern = r'SG_\s+([^:]+)\s*:\s*(\d+)\|(\d+)@([01])([+-])\s*\(([^,]+),([^)]+)\)\s*\[([^|]+)\|([^\]]+)\]\s*"([^"]*)"\s*([^\n]*)'
        
        current_msg_id = None
        
        # Dividir el contenido en líneas para procesamiento secuencial
        lines = self.dbc_content.split('\n')
        
        for line in lines:
            # Detectar inicio de mensaje
            msg_match = re.match(r'BO_\s+(\d+)', line)
            if msg_match:
                current_msg_id = int(msg_match.group(1))
                continue
            
            # Procesar señales
            signal_match = re.match(signal_pattern, line.strip())
            if signal_match and current_msg_id is not None:
                signal_name = signal_match.group(1).strip()
                start_bit = int(signal_match.group(2))
                length = int(signal_match.group(3))
                endianness = 'Little Endian' if signal_match.group(4) == '1' else 'Big Endian'
                sign = 'Signed' if signal_match.group(5) == '-' else 'Unsigned'
                factor = float(signal_match.group(6))
                offset = float(signal_match.group(7))
                min_val = float(signal_match.group(8))
                max_val = float(signal_match.group(9))
                unit = signal_match.group(10)
                receivers = signal_match.group(11).strip()
                
                signal_info = {
                    'name': signal_name,
                    'message_id': current_msg_id,
                    'message_name': self.messages.get(current_msg_id, {}).get('name', ''),
                    'start_bit': start_bit,
                    'length': length,
                    'endianness': endianness,
                    'sign': sign,
                    'factor': factor,
                    'offset': offset,
                    'min_value': min_val,
                    'max_value': max_val,
                    'unit': unit,
                    'receivers': receivers
                }
                
                # Agregar señal al diccionario global
                self.signals[f"{current_msg_id}_{signal_name}"] = signal_info
                
                # Agregar señal al mensaje correspondiente
                if current_msg_id in self.messages:
                    self.messages[current_msg_id]['signals'].append(signal_info)
        
        logger.info(f"Señales encontradas: {len(self.signals)}")
    
    def parse_value_tables(self):
        """
        Extrae las tablas de valores del archivo DBC
        """
        # Patrón para tablas de valores: VAL_ <MessageID> <SignalName> <Value> "<Text>" ;
        value_pattern = r'VAL_\s+(\d+)\s+([^\s]+)\s+(.+?);'
        
        for match in re.finditer(value_pattern, self.dbc_content, re.DOTALL):
            msg_id = int(match.group(1))
            signal_name = match.group(2).strip()
            values_str = match.group(3).strip()
            
            # Parsear los valores
            value_pairs = re.findall(r'(\d+)\s+"([^"]+)"', values_str)
            
            key = f"{msg_id}_{signal_name}"
            self.value_tables[key] = {}
            
            for value, text in value_pairs:
                self.value_tables[key][int(value)] = text
        
        logger.info(f"Tablas de valores encontradas: {len(self.value_tables)}")
    
    def parse_attributes(self):
        """
        Extrae los atributos del archivo DBC
        """
        # Patrón para definiciones de atributos: BA_DEF_ "<Name>" <Type> [<Options>];
        attr_def_pattern = r'BA_DEF_\s+"([^"]+)"\s+([^;]+);'
        
        # Patrón para valores de atributos: BA_ "<Name>" [<Node>|<Message>|<Signal>] <Value>;
        attr_val_pattern = r'BA_\s+"([^"]+)"\s*(.*?)\s*([^;]+);'
        
        # Parsear definiciones de atributos
        for match in re.finditer(attr_def_pattern, self.dbc_content):
            attr_name = match.group(1)
            attr_type = match.group(2).strip()
            
            if attr_name not in self.attributes:
                self.attributes[attr_name] = {'type': attr_type, 'values': {}}
        
        # Parsear valores de atributos
        for match in re.finditer(attr_val_pattern, self.dbc_content):
            attr_name = match.group(1)
            target = match.group(2).strip()
            value = match.group(3).strip()
            
            if attr_name not in self.attributes:
                self.attributes[attr_name] = {'type': 'unknown', 'values': {}}
            
            self.attributes[attr_name]['values'][target if target else 'global'] = value
        
        logger.info(f"Atributos encontrados: {len(self.attributes)}")
    
    def parse_comments(self):
        """
        Extrae los comentarios del archivo DBC
        """
        # Patrón para comentarios: CM_ [<Type> <ID>] "<Comment>";
        comment_pattern = r'CM_\s*(.*?)\s*"([^"]+)";'
        
        for match in re.finditer(comment_pattern, self.dbc_content):
            target = match.group(1).strip()
            comment = match.group(2)
            
            if target:
                self.comments[target] = comment
            else:
                self.comments['general'] = comment
        
        logger.info(f"Comentarios encontrados: {len(self.comments)}")
    
    def process_dbc(self) -> bool:
        """
        Procesa completamente el archivo DBC
        
        Returns:
            bool: True si el procesamiento fue exitoso
        """
        if not self.load_dbc_file():
            return False
        
        logger.info("Iniciando procesamiento del archivo DBC...")
        
        self.parse_nodes()
        self.parse_messages()
        self.parse_signals()
        self.parse_value_tables()
        self.parse_attributes()
        self.parse_comments()
        
        logger.info("Procesamiento del archivo DBC completado")
        return True
    
    def get_messages_dataframe(self) -> pd.DataFrame:
        """
        Convierte los mensajes a un DataFrame de pandas
        
        Returns:
            pd.DataFrame: DataFrame con información de mensajes
        """
        messages_data = []
        
        for msg_id, msg_info in self.messages.items():
            messages_data.append({
                'ID': msg_info['id'],
                'ID_Hex': f"0x{msg_info['id']:X}",
                'Nombre': msg_info['name'],
                'Tamaño (bytes)': msg_info['size'],
                'Nodo': msg_info['node'],
                'Num_Señales': len(msg_info['signals'])
            })
        
        return pd.DataFrame(messages_data)
    
    def get_signals_dataframe(self) -> pd.DataFrame:
        """
        Convierte las señales a un DataFrame de pandas
        
        Returns:
            pd.DataFrame: DataFrame con información de señales
        """
        signals_data = []
        
        for signal_key, signal_info in self.signals.items():
            # Obtener tabla de valores si existe
            value_table = self.value_tables.get(signal_key, {})
            value_table_str = '; '.join([f"{k}:{v}" for k, v in value_table.items()]) if value_table else ""
            
            signals_data.append({
                'Mensaje_ID': signal_info['message_id'],
                'Mensaje_ID_Hex': f"0x{signal_info['message_id']:X}",
                'Mensaje_Nombre': signal_info['message_name'],
                'Señal_Nombre': signal_info['name'],
                'Bit_Inicio': signal_info['start_bit'],
                'Longitud': signal_info['length'],
                'Endianness': signal_info['endianness'],
                'Signo': signal_info['sign'],
                'Factor': signal_info['factor'],
                'Offset': signal_info['offset'],
                'Valor_Min': signal_info['min_value'],
                'Valor_Max': signal_info['max_value'],
                'Unidad': signal_info['unit'],
                'Receptores': signal_info['receivers'],
                'Tabla_Valores': value_table_str
            })
        
        return pd.DataFrame(signals_data)
    
    def get_nodes_dataframe(self) -> pd.DataFrame:
        """
        Convierte los nodos a un DataFrame de pandas
        
        Returns:
            pd.DataFrame: DataFrame con información de nodos
        """
        nodes_data = []
        
        for i, node in enumerate(self.nodes):
            # Contar mensajes por nodo
            msg_count = sum(1 for msg in self.messages.values() if msg['node'] == node)
            
            nodes_data.append({
                'Index': i + 1,
                'Nombre': node,
                'Num_Mensajes': msg_count
            })
        
        return pd.DataFrame(nodes_data)
    
    def get_attributes_dataframe(self) -> pd.DataFrame:
        """
        Convierte los atributos a un DataFrame de pandas
        
        Returns:
            pd.DataFrame: DataFrame con información de atributos
        """
        attributes_data = []
        
        for attr_name, attr_info in self.attributes.items():
            for target, value in attr_info['values'].items():
                attributes_data.append({
                    'Atributo': attr_name,
                    'Tipo': attr_info['type'],
                    'Objetivo': target,
                    'Valor': value
                })
        
        return pd.DataFrame(attributes_data)
    
    def get_value_tables_dataframe(self) -> pd.DataFrame:
        """
        Convierte las tablas de valores a un DataFrame de pandas
        
        Returns:
            pd.DataFrame: DataFrame con información de tablas de valores
        """
        value_tables_data = []
        
        for key, value_table in self.value_tables.items():
            msg_id, signal_name = key.split('_', 1)
            
            for value, text in value_table.items():
                value_tables_data.append({
                    'Mensaje_ID': int(msg_id),
                    'Mensaje_ID_Hex': f"0x{int(msg_id):X}",
                    'Señal': signal_name,
                    'Valor': value,
                    'Texto': text
                })
        
        return pd.DataFrame(value_tables_data)
    
    def get_comments_dataframe(self) -> pd.DataFrame:
        """
        Convierte los comentarios a un DataFrame de pandas
        
        Returns:
            pd.DataFrame: DataFrame con información de comentarios
        """
        comments_data = []
        
        for target, comment in self.comments.items():
            comments_data.append({
                'Objetivo': target,
                'Comentario': comment
            })
        
        return pd.DataFrame(comments_data)


class ExcelExporter:
    """
    Clase para exportar datos DBC a archivo Excel con formato
    """
    
    def __init__(self, output_file: str):
        """
        Inicializa el exportador Excel
        
        Args:
            output_file (str): Ruta del archivo Excel de salida
        """
        self.output_file = output_file
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remover hoja por defecto
    
    def _apply_header_style(self, worksheet, row_num: int):
        """
        Aplica estilo a los encabezados
        
        Args:
            worksheet: Hoja de trabajo de Excel
            row_num (int): Número de fila del encabezado
        """
        for cell in worksheet[row_num]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
    
    def _apply_data_style(self, worksheet):
        """
        Aplica estilo a los datos
        
        Args:
            worksheet: Hoja de trabajo de Excel
        """
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                cell.alignment = Alignment(vertical="center")
    
    def _auto_adjust_columns(self, worksheet):
        """
        Ajusta automáticamente el ancho de las columnas
        
        Args:
            worksheet: Hoja de trabajo de Excel
        """
        # Obtener dimensiones de la hoja
        if worksheet.max_column == 0:
            return
            
        for col_num in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            
            # Revisar todas las celdas de la columna
            for row_num in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max(max_length + 2, 10), 50)  # Mínimo 10, máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_dataframe_sheet(self, df: pd.DataFrame, sheet_name: str, description: str = ""):
        """
        Agrega una hoja de Excel con un DataFrame
        
        Args:
            df (pd.DataFrame): DataFrame a exportar
            sheet_name (str): Nombre de la hoja
            description (str): Descripción opcional
        """
        if df.empty:
            logger.warning(f"DataFrame vacío para hoja: {sheet_name}")
            # Crear hoja vacía con mensaje
            worksheet = self.workbook.create_sheet(title=sheet_name)
            worksheet.cell(row=1, column=1, value=f"No se encontraron datos para: {sheet_name}")
            worksheet.cell(row=1, column=1).font = Font(bold=True, color="FF0000")
            return
        
        worksheet = self.workbook.create_sheet(title=sheet_name)
        
        # Agregar descripción si se proporciona
        if description:
            worksheet.cell(row=1, column=1, value=description)
            worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            start_row = 3
        else:
            start_row = 1
        
        # Agregar datos del DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Aplicar estilos
        self._apply_header_style(worksheet, start_row)
        self._apply_data_style(worksheet)
        self._auto_adjust_columns(worksheet)
        
        logger.info(f"Hoja '{sheet_name}' agregada con {len(df)} filas")
    
    def add_summary_sheet(self, dbc_processor: DBCProcessor):
        """
        Agrega una hoja de resumen con estadísticas del archivo DBC
        
        Args:
            dbc_processor (DBCProcessor): Instancia del procesador DBC
        """
        worksheet = self.workbook.create_sheet(title="Resumen", index=0)
        
        # Título
        worksheet.cell(row=1, column=1, value="RESUMEN DEL ARCHIVO DBC")
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=16)
        worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        
        # Información del archivo
        info_data = [
            ("Archivo:", os.path.basename(dbc_processor.dbc_file_path)),
            ("Ruta completa:", dbc_processor.dbc_file_path),
            ("", ""),
            ("ESTADÍSTICAS:", ""),
            ("Total de nodos:", len(dbc_processor.nodes)),
            ("Total de mensajes:", len(dbc_processor.messages)),
            ("Total de señales:", len(dbc_processor.signals)),
            ("Total de tablas de valores:", len(dbc_processor.value_tables)),
            ("Total de atributos:", len(dbc_processor.attributes)),
            ("Total de comentarios:", len(dbc_processor.comments))
        ]
        
        for row_idx, (label, value) in enumerate(info_data, start=3):
            worksheet.cell(row=row_idx, column=1, value=label)
            worksheet.cell(row=row_idx, column=2, value=value)
            
            if label.endswith(":") and not label.startswith("ESTADÍSTICAS"):
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Aplicar ajustar columnas
        self._auto_adjust_columns(worksheet)
        
        logger.info("Hoja de resumen agregada")
    
    def save(self):
        """
        Guarda el archivo Excel
        """
        try:
            self.workbook.save(self.output_file)
            logger.info(f"Archivo Excel guardado exitosamente: {self.output_file}")
        except Exception as e:
            logger.error(f"Error al guardar archivo Excel: {e}")
            raise


def process_dbc_to_excel(dbc_file_path: str, excel_output_path: str) -> bool:
    """
    Función principal para procesar archivo DBC y exportar a Excel
    
    Args:
        dbc_file_path (str): Ruta al archivo DBC
        excel_output_path (str): Ruta del archivo Excel de salida
    
    Returns:
        bool: True si el proceso fue exitoso
    """
    try:
        logger.info(f"Iniciando procesamiento: {dbc_file_path} -> {excel_output_path}")
        
        # Procesar archivo DBC
        processor = DBCProcessor(dbc_file_path)
        if not processor.process_dbc():
            logger.error("Error en el procesamiento del archivo DBC")
            return False
        
        # Crear exportador Excel
        exporter = ExcelExporter(excel_output_path)
        
        # Agregar hoja de resumen
        exporter.add_summary_sheet(processor)
        
        # Agregar hojas con datos
        exporter.add_dataframe_sheet(
            processor.get_messages_dataframe(),
            "Mensajes",
            "Información detallada de todos los mensajes CAN en el archivo DBC"
        )
        
        exporter.add_dataframe_sheet(
            processor.get_signals_dataframe(),
            "Señales",
            "Información detallada de todas las señales CAN en el archivo DBC"
        )
        
        exporter.add_dataframe_sheet(
            processor.get_nodes_dataframe(),
            "Nodos",
            "Lista de todos los nodos definidos en el archivo DBC"
        )
        
        if processor.value_tables:
            exporter.add_dataframe_sheet(
                processor.get_value_tables_dataframe(),
                "Tablas_Valores",
                "Tablas de valores para señales específicas"
            )
        
        if processor.attributes:
            exporter.add_dataframe_sheet(
                processor.get_attributes_dataframe(),
                "Atributos",
                "Atributos definidos en el archivo DBC"
            )
        
        if processor.comments:
            exporter.add_dataframe_sheet(
                processor.get_comments_dataframe(),
                "Comentarios",
                "Comentarios incluidos en el archivo DBC"
            )
        
        # Guardar archivo
        exporter.save()
        
        logger.info("Procesamiento completado exitosamente")
        return True
        
    except Exception as e:
        logger.error(f"Error durante el procesamiento: {e}")
        return False